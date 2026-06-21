import urllib.request
import openpyxl
import unicodedata

def normalize_team_name(name):
    if name is None:
        return ""
    s = str(name).strip().lower()
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = s.replace(".", "").replace("-", " ").replace("  ", " ")
    return s

TRADUCCIONES_EQUIPOS = {
    "Mexico": "México", "South Korea": "Corea del Sur", "South Africa": "Sudáfrica", "Czechia": "Chequia", "Czech Republic": "Chequia",
    "Canada": "Canadá", "Switzerland": "Suiza", "Qatar": "Qatar", "Bosnia and Herzegovina": "Bosnia y Herze.", "Bosnia & Herzegovina": "Bosnia y Herze.",
    "Brazil": "Brasil", "Scotland": "Escocia", "Morocco": "Marruecos", "Haiti": "Haití",
    "USA": "Estados Unidos", "United States": "Estados Unidos", "Turkey": "Turquía", "Türkiye": "Turquía", "Paraguay": "Paraguay", "Australia": "Australia",
    "Germany": "Alemania", "Ecuador": "Ecuador", "Ivory Coast": "Costa de Marfil", "Côte d'Ivoire": "Costa de Marfil", "Curacao": "Curaçao",
    "Netherlands": "Países Bajos", "Tunisia": "Túnez", "Japan": "Japón", "Sweden": "Suecia",
    "Belgium": "Bélgica", "New Zealand": "Nueva Zelanda", "Egypt": "Egipto", "Iran": "Irán", "IR Iran": "Irán",
    "Spain": "España", "Uruguay": "Uruguay", "Cape Verde": "Cabo Verde", "Saudi Arabia": "Arabia Saud.",
    "France": "Francia", "Norway": "Noruega", "Senegal": "Senegal", "Iraq": "Irak",
    "Argentina": "Argentina", "Jordan": "Jordania", "Algeria": "Argelia", "Austria": "Austria",
    "Portugal": "Portugal", "Colombia": "Colombia", "DR Congo": "RD Congo", "Congo DR": "RD Congo", "Uzbekistan": "Uzbekistán",
    "England": "Inglaterra", "Panama": "Panamá", "Croatia": "Croacia", "Ghana": "Ghana"
}

def traducir(nombre):
    return TRADUCCIONES_EQUIPOS.get(nombre, nombre)

def update_master_from_web(template_path, output_path):
    import time
    import json
    import urllib.request
    url = f"https://raw.githubusercontent.com/openfootball/worldcup.json/master/2026/worldcup.json?_t={int(time.time())}"
    
    # Descargar JSON
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        response = urllib.request.urlopen(req, timeout=10)
        json_data = response.read().decode('utf-8')
    except Exception as e:
        raise Exception(f"No se pudo descargar los datos de internet: {e}")
        
    try:
        data = json.loads(json_data)
    except json.JSONDecodeError:
        raise Exception("El archivo descargado no tiene un formato JSON válido.")
        
    if "matches" not in data:
        raise Exception("El archivo JSON no contiene la lista de partidos.")
        
    # Parsear JSON en un diccionario para búsquedas rápidas
    # Clave: (norm_equipo_local, norm_equipo_visitante), Valor: (goles_L, goles_V)
    web_scores = {}
    for match in data.get("matches", []):
        if "score" in match and "ft" in match["score"]:
            score_ft = match["score"]["ft"]
            if len(score_ft) == 2:
                home_en = match.get("team1", "")
                away_en = match.get("team2", "")
                score_h = score_ft[0]
                score_a = score_ft[1]
                
                home_es = traducir(home_en)
                away_es = traducir(away_en)
                
                norm_h = normalize_team_name(home_es)
                norm_a = normalize_team_name(away_es)
                
                web_scores[(norm_h, norm_a)] = (score_h, score_a)
            
    # Cargar plantilla de Excel
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        raise Exception(f"No se pudo abrir la plantilla {template_path}: {e}")
        
    ws = wb['Grupos']
    
    # Recorrer todos los partidos de grupos en la plantilla
    partidos_actualizados = 0
    for g in range(12):
        start_row = 6 + (8 * g)
        for m in range(6):
            r = start_row + m
            t1 = ws.cell(row=r, column=3).value
            t2 = ws.cell(row=r, column=7).value
            
            if t1 and t2:
                norm_t1 = normalize_team_name(t1)
                norm_t2 = normalize_team_name(t2)
                
                # Buscar si el partido existe en web_scores
                # Puede estar invertido (t1 como visitante en la web)
                if (norm_t1, norm_t2) in web_scores:
                    s1, s2 = web_scores[(norm_t1, norm_t2)]
                    ws.cell(row=r, column=4, value=s1)
                    ws.cell(row=r, column=6, value=s2)
                    partidos_actualizados += 1
                elif (norm_t2, norm_t1) in web_scores:
                    s2, s1 = web_scores[(norm_t2, norm_t1)]
                    ws.cell(row=r, column=4, value=s1)
                    ws.cell(row=r, column=6, value=s2)
                    partidos_actualizados += 1
                    
    # Guardar archivo
    wb.save(output_path)
    return partidos_actualizados
