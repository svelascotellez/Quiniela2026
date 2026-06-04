#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
fill_quiniela.py
Llena Quiniela_Mundial_2026_Final_vacia (1).xlsx usando:
  - Ranking FIFA (leído desde la misma hoja del archivo)
  - Probabilidades Opta/The Analyst (World Cup 2026)
  - Factor de aleatoriedad

Uso:
  C:\\Python313\\python.exe fill_quiniela.py
"""

import random
import openpyxl

# ─── SEMILLA ALEATORIA ─────────────────────────────────────────────
random.seed(42)   # Cambia para obtener un resultado diferente

# ─── ESTRUCTURA DE LA HOJA GRUPOS ──────────────────────────────────
# Columna C(3) = Equipo 1, D(4) = Goles 1, F(6) = Goles 2, G(7) = Equipo 2
# Grupos: A→6-11, B→14-19, C→22-27, D→30-35, E→38-43, F→46-51,
#         G→54-59, H→62-67, I→70-75, J→78-83, K→86-91, L→94-99
GROUP_ROWS = {
    "A": list(range(6,  12)),
    "B": list(range(14, 20)),
    "C": list(range(22, 28)),
    "D": list(range(30, 36)),
    "E": list(range(38, 44)),
    "F": list(range(46, 52)),
    "G": list(range(54, 60)),
    "H": list(range(62, 68)),
    "I": list(range(70, 76)),
    "J": list(range(78, 84)),
    "K": list(range(86, 92)),
    "L": list(range(94, 100)),
}

# ─── ESTRUCTURA ELIMINATORIAS ───────────────────────────────────────
# Columna B(2)=Equipo A (fórmula), C(3)=Goles A, D(4)=Pen A,
# F(6)=Pen B, G(7)=Goles B, H(8)=Equipo B (fórmula), I(9)=Clasifica (fórmula)
R32_ROWS   = list(range(6,  22))   # P1-P16
R16_ROWS   = list(range(25, 33))   # 8 octavos
QF_ROWS    = list(range(37, 41))   # 4 cuartos
SF_ROWS    = [45, 46]              # 2 semifinales
THIRD_ROW  = 51
FINAL_ROW  = 56

# ─── PROBABILIDADES OPTA (% ganador del torneo, junio 2026) ────────
OPTA_WIN_PROB = {
    "España":          16.12,
    "Francia":         12.98,
    "Inglaterra":      11.18,
    "Argentina":       10.36,
    "Portugal":         7.00,
    "Brasil":           6.61,
    "Alemania":         5.12,
    "Países Bajos":     3.80,
    "Marruecos":        3.50,
    "Bélgica":          2.90,
    "Estados Unidos":   2.70,
    "Croacia":          2.20,
    "México":           2.00,
    "Uruguay":          1.80,
    "Colombia":         1.70,
    "Japón":            1.60,
    "Turquía":          1.50,
    "Suiza":            1.40,
    "Senegal":          1.20,
    "Ecuador":          1.10,
    "Austria":          1.10,
    "Corea del Sur":    1.00,
    "Australia":        0.85,
    "Dinamarca":        0.80,
    "Suecia":           0.70,
    "Canadá":           0.65,
    "Ghana":            0.50,
    "Serbia":           0.50,
    "Costa de Marfil":  0.60,
    "Chequia":          0.50,
    "Polonia":          0.45,
    "Noruega":          0.45,
    "Paraguay":         0.40,
    "Argelia":          0.40,
    "Hungría":          0.35,
    "Egipto":           0.35,
    "Escocia":          0.30,
    "Túnez":            0.30,
    "Camerún":          0.25,
    "Ucrania":          0.25,
    "Bosnia y Herze.":  0.25,
    "Arabia Saud.":     0.20,
    "Uzbekistán":       0.20,
    "Haití":            0.10,
    "Irak":             0.15,
    "Qatar":            0.15,
    "Sudáfrica":        0.15,
    "RD Congo":         0.15,
    "Jordania":         0.12,
    "Cabo Verde":       0.10,
    "Nueva Zelanda":    0.10,
    "Irán":             0.15,
    "Panamá":           0.40,
    "Curaçao":          0.05,
}

def opta_strength(team: str) -> float:
    return OPTA_WIN_PROB.get(str(team).strip(), 0.50)

def fifa_rank(team: str, ranks: dict) -> int:
    return ranks.get(str(team).strip(), 100)

def combined_strength(team: str, ranks: dict) -> float:
    rank = fifa_rank(team, ranks)
    opta = opta_strength(team)
    return (212 - rank) * (1 + opta / 8)

def simulate_score(team_a: str, team_b: str, ranks: dict,
                   noise: float = 0.35, knockout: bool = False):
    """
    Simula un marcador realista.
    Retorna (g_a, g_b, pen_a, pen_b).
    pen_* = None si no hay penales.
    """
    sa = combined_strength(team_a, ranks) * random.uniform(1 - noise, 1 + noise)
    sb = combined_strength(team_b, ranks) * random.uniform(1 - noise, 1 + noise)
    ratio = sa / (sa + sb)   # prob implícita de victoria de A

    r = random.random()

    # Marcadores típicos del fútbol (distribución ajustada)
    def win_scorelines(winner_goals, loser_goals_choices):
        return random.choice(winner_goals), random.choice(loser_goals_choices)

    if r < ratio * 0.55:
        g_a, g_b = win_scorelines([2, 3, 3, 4], [0, 0, 1])
    elif r < ratio * 0.80:
        g_a, g_b = win_scorelines([1, 2], [0, 0, 1])
        if g_a == g_b:
            g_a += 1
    elif r < 0.5 and not knockout:
        base = random.choice([0, 1, 1, 2])
        g_a = g_b = base
    elif r > (1 - ratio) * 0.20 + 0.5:
        g_b, g_a = win_scorelines([1, 2], [0, 0, 1])
        if g_a == g_b:
            g_b += 1
    else:
        g_b, g_a = win_scorelines([2, 3, 3, 4], [0, 0, 1])

    # Empate en knockout → penales
    pen_a = pen_b = None
    if knockout and g_a == g_b:
        if random.random() < ratio:
            pen_a = random.choice([4, 5, 5, 6])
            pen_b = pen_a - random.choice([1, 2])
        else:
            pen_b = random.choice([4, 5, 5, 6])
            pen_a = pen_b - random.choice([1, 2])

    return g_a, g_b, pen_a, pen_b


# ─── LEER RANKING FIFA ─────────────────────────────────────────────
def load_fifa_ranks(wb) -> dict:
    ws = wb["Ranking FIFA"]
    ranks = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            try:
                ranks[str(row[1]).strip()] = int(row[0])
            except (ValueError, TypeError):
                pass
    return ranks


# ─── FASE DE GRUPOS ────────────────────────────────────────────────
def fill_group_stage(ws, ranks: dict):
    print("\n[GRUPOS]")
    for letter, rows in GROUP_ROWS.items():
        for row in rows:
            t1 = ws.cell(row=row, column=3).value
            t2 = ws.cell(row=row, column=7).value
            g1 = ws.cell(row=row, column=4).value
            g2 = ws.cell(row=row, column=6).value

            if t1 and t2 and (g1 is None or g2 is None):
                sg1, sg2, _, _ = simulate_score(t1, t2, ranks, noise=0.38)
                ws.cell(row=row, column=4).value = sg1
                ws.cell(row=row, column=6).value = sg2
                print(f"  Gr.{letter} R{row:03d}: {t1} {sg1}-{sg2} {t2}")


# ─── CALCULAR CLASIFICADOS DE GRUPOS ──────────────────────────────
def compute_group_standings(ws, ranks: dict) -> dict:
    standings = {}
    for letter, rows in GROUP_ROWS.items():
        teams = {}
        for row in rows:
            t1 = str(ws.cell(row=row, column=3).value or "").strip()
            t2 = str(ws.cell(row=row, column=7).value or "").strip()
            g1 = ws.cell(row=row, column=4).value
            g2 = ws.cell(row=row, column=6).value
            for t in [t1, t2]:
                if t and t not in teams:
                    teams[t] = {"pts": 0, "gf": 0, "gc": 0, "gd": 0}
            if t1 and t2 and g1 is not None and g2 is not None:
                g1, g2 = int(g1), int(g2)
                teams[t1]["gf"] += g1; teams[t1]["gc"] += g2; teams[t1]["gd"] += g1 - g2
                teams[t2]["gf"] += g2; teams[t2]["gc"] += g1; teams[t2]["gd"] += g2 - g1
                if g1 > g2: teams[t1]["pts"] += 3
                elif g1 < g2: teams[t2]["pts"] += 3
                else: teams[t1]["pts"] += 1; teams[t2]["pts"] += 1

        sorted_teams = sorted(
            teams.keys(),
            key=lambda t: (teams[t]["pts"], teams[t]["gd"], teams[t]["gf"],
                           -(fifa_rank(t, ranks))),
            reverse=True
        )
        standings[letter] = [{"team": t, **teams[t]} for t in sorted_teams]
    return standings


def get_best_thirds(standings: dict) -> list:
    thirds = []
    for letter, s in standings.items():
        if len(s) >= 3:
            thirds.append({"group": letter, **s[2]})
    thirds.sort(key=lambda x: (x["pts"], x["gd"], x["gf"]), reverse=True)
    return thirds[:8]


# ─── LEER EQUIPOS DE ELIMINATORIAS (desde fórmulas del Backend) ───
# Los equipos A y B en la hoja Eliminatorias provienen de fórmulas del Backend
# que NO se evalúan con data_only=False. En su lugar, reconstruimos los emparejamientos
# desde los standings de grupo siguiendo el esquema oficial del Backend.
def build_r32_teams(standings: dict, thirds: list):
    """
    Mapeo oficial de los dieciseisavos según las fórmulas del Backend.
    Backend!B<n> = 1ro del grupo, Backend!C<n> = 2do del grupo
    Fila Backend: A=2, B=3, C=4, D=5, E=6, F=7, G=8, H=9, I=10, J=11, K=12, L=13

    Emparejamientos (de inspect_eliminatorias.py):
      R6  P1:  C2=2do_A   vs C3=2do_B?   → Backend!C2 = 2do_A, Backend!C3 = 2do_B
           (Revisado: C2 = 2do_A, C3 = 1ro_B en muchos esquemas; usamos el oficial)
      R7  P2:  B7=1ro_F   vs C4=2do_C
      R8  P3:  B6=1ro_E   vs mejor_3ro_1
      R9  P4:  B10=1ro_J  vs mejor_3ro_2
      R10 P5:  B4=1ro_D   vs C7=2do_F
      R11 P6:  C6=2do_E   vs C10=2do_J
      R12 P7:  B2=1ro_A   vs mejor_3ro_3
      R13 P8:  B13=1ro_L  vs mejor_3ro_4
      R14 P9:  B5=1ro_E?  vs mejor_3ro_5   (B5=1ro_D en notación Backend fila 5)
               → B5 = 1ro grupo D? No: Backend fila 5 = Grupo D, B5=1ro_D (conflicto con P5)
               → Reinterpretando: Backend!B5 podría ser 1ro_D, pero P5 ya usa 1ro_D.
               → En la quiniela: P9: B5=1ro_C (Backend fila 4=C→ 1ro_C está en B4? No...)
               → Backend rows: A=2,B=3,C=4,D=5,E=6,F=7,G=8,H=9,I=10,J=11,K=12,L=13
               → B5=Backend!B5=1ro_D, C4=Backend!C4=2do_C
               → Para evitar conflicto P5 vs P9 usando mismo 1ro_D:
                 P9 en realidad es Backend!B5=1ro_D? El archivo los separa en P5(1ro_D vs 2do_F)
                 vs P9(1ro_D otra vez)? Revisando el output:
                 R010: 1/16-P5: =Backend!B4  vs =Backend!C7 → B4=1ro_C? No: fila 4=Grupo C
                    Espera: Backend filas A=2,B=3,C=4,D=5 → B4 = 1ro grupo C (fila 4), C4=2do_C
                    Backend B5=1ro_D, C5=2do_D, B6=1ro_E, C6=2do_E...
               → Corrección: R10-P5: B4=1ro_C (¡no D!), C7=2do_F
               → R14-P9: B5=1ro_D

    Mapeo final corregido (Backend fila = grupo: A=2,B=3,C=4,D=5,E=6,F=7,G=8,H=9,I=10,J=11,K=12,L=13):
      P1:  C2=2do_A   vs C3=2do_B         → Posicion (A,1) vs (B,1)
      P2:  B7=1ro_F   vs C4=2do_C         → (F,0) vs (C,1)
      P3:  B6=1ro_E   vs 3ro_best[0]
      P4:  B10=1ro_I  vs 3ro_best[1]       (fila 10=I)
      P5:  B4=1ro_C   vs C7=2do_F         → (C,0) vs (F,1)
      P6:  C6=2do_E   vs C10=2do_I        → (E,1) vs (I,1)
      P7:  B2=1ro_A   vs 3ro_best[2]
      P8:  B13=1ro_L  vs 3ro_best[3]
      P9:  B5=1ro_D   vs 3ro_best[4]
      P10: B8=1ro_G   vs 3ro_best[5]
      P11: C12=2do_K  vs C13=2do_L        → (K,1) vs (L,1)
      P12: B9=1ro_H   vs C11=2do_J        → (H,0) vs (J,1)
      P13: B3=1ro_B   vs 3ro_best[6]
      P14: B12=1ro_K  vs 3ro_best[7]
      P15: B11=1ro_J  vs C9=2do_H         → (J,0) vs (H,1)
      P16: C5=2do_D   vs C8=2do_G         → (D,1) vs (G,1)
    """

    def team(g, pos):
        s = standings.get(g, [])
        return s[pos]["team"] if pos < len(s) else "???"

    t3 = [x["team"] for x in thirds]  # 8 mejores terceros
    while len(t3) < 8:
        t3.append("???")

    return [
        # row, team_a, team_b
        (R32_ROWS[0],  team("A",1), team("B",1)),  # P1
        (R32_ROWS[1],  team("F",0), team("C",1)),  # P2
        (R32_ROWS[2],  team("E",0), t3[0]),         # P3
        (R32_ROWS[3],  team("I",0), t3[1]),         # P4
        (R32_ROWS[4],  team("C",0), team("F",1)),  # P5
        (R32_ROWS[5],  team("E",1), team("I",1)),  # P6
        (R32_ROWS[6],  team("A",0), t3[2]),         # P7
        (R32_ROWS[7],  team("L",0), t3[3]),         # P8
        (R32_ROWS[8],  team("D",0), t3[4]),         # P9
        (R32_ROWS[9],  team("G",0), t3[5]),         # P10
        (R32_ROWS[10], team("K",1), team("L",1)),  # P11
        (R32_ROWS[11], team("H",0), team("J",1)),  # P12
        (R32_ROWS[12], team("B",0), t3[6]),         # P13
        (R32_ROWS[13], team("K",0), t3[7]),         # P14
        (R32_ROWS[14], team("J",0), team("H",1)),  # P15
        (R32_ROWS[15], team("D",1), team("G",1)),  # P16
    ]


# ─── SIMULAR Y ESCRIBIR MARCADORES ─────────────────────────────────
def sim_and_write(ws, row, team_a, team_b, ranks, noise, col_ga=3, col_pen_a=4,
                  col_pen_b=6, col_gb=7, knockout=True):
    ga, gb, pen_a, pen_b = simulate_score(team_a, team_b, ranks, noise, knockout)
    ws.cell(row=row, column=col_ga).value   = ga
    ws.cell(row=row, column=col_gb).value   = gb
    if pen_a is not None:
        ws.cell(row=row, column=col_pen_a).value = pen_a
        ws.cell(row=row, column=col_pen_b).value = pen_b

    if knockout:
        if ga > gb or (ga == gb and pen_a is not None and pen_a > pen_b):
            winner = team_a
        else:
            winner = team_b
    else:
        winner = None

    pen_str = f" ({pen_a}p-{pen_b}p)" if pen_a is not None else ""
    print(f"    {team_a} {ga}-{gb} {team_b}{pen_str}"
          + (f"  -> {winner}" if winner else ""))
    return winner


# ─── FASE ELIMINATORIA COMPLETA ────────────────────────────────────
def fill_knockout_stage(ws, standings, thirds, ranks):
    r32_matchups = build_r32_teams(standings, thirds)

    print("\n[DIECISEISAVOS]")
    r32_winners = []
    for row, ta, tb in r32_matchups:
        w = sim_and_write(ws, row, ta, tb, ranks, noise=0.30)
        r32_winners.append(w)

    # Octavos: pares consecutivos de R32 (P1vsP2, P3vsP4, ...)
    print("\n[OCTAVOS]")
    r16_winners = []
    for i, row in enumerate(R16_ROWS):
        ta = r32_winners[i * 2]
        tb = r32_winners[i * 2 + 1]
        w = sim_and_write(ws, row, ta, tb, ranks, noise=0.28)
        r16_winners.append(w)

    # Cuartos
    print("\n[CUARTOS]")
    qf_winners = []
    for i, row in enumerate(QF_ROWS):
        ta = r16_winners[i * 2]
        tb = r16_winners[i * 2 + 1]
        w = sim_and_write(ws, row, ta, tb, ranks, noise=0.25)
        qf_winners.append(w)

    # Semifinales
    print("\n[SEMIFINALES]")
    sf_winners = []
    sf_losers  = []
    for i, row in enumerate(SF_ROWS):
        ta = qf_winners[i * 2]
        tb = qf_winners[i * 2 + 1]
        ga = ws.cell(row=row, column=3)
        gb = ws.cell(row=row, column=7)
        ga2, gb2, pen_a, pen_b = simulate_score(ta, tb, ranks, noise=0.22, knockout=True)
        ws.cell(row=row, column=3).value = ga2
        ws.cell(row=row, column=7).value = gb2
        if pen_a is not None:
            ws.cell(row=row, column=4).value = pen_a
            ws.cell(row=row, column=6).value = pen_b

        if ga2 > gb2 or (ga2 == gb2 and pen_a is not None and pen_a > pen_b):
            w, l = ta, tb
        else:
            w, l = tb, ta

        pen_str = f" ({pen_a}p-{pen_b}p)" if pen_a is not None else ""
        print(f"    {ta} {ga2}-{gb2} {tb}{pen_str}  -> {w}")
        sf_winners.append(w)
        sf_losers.append(l)

    # Tercer lugar
    print("\n[TERCER LUGAR]")
    third = sim_and_write(ws, THIRD_ROW, sf_losers[0], sf_losers[1], ranks, noise=0.28)

    # Gran Final
    print("\n[GRAN FINAL]")
    champion = sim_and_write(ws, FINAL_ROW, sf_winners[0], sf_winners[1], ranks, noise=0.20)

    return champion, sf_winners[1] if champion == sf_winners[0] else sf_winners[0], third


# ─── MAIN ──────────────────────────────────────────────────────────
def main():
    INPUT  = r"C:\Users\salva\Downloads\Quiniela_Mundial_2026_Final_IA.xlsx"
    OUTPUT = r"C:\Users\salva\.antigravity\Quiniela2026\Quiniela_Mundial_2026_IA_Llena.xlsx"

    print("Cargando archivo...")
    wb = openpyxl.load_workbook(INPUT, data_only=False)

    print("Leyendo Ranking FIFA...")
    ranks = load_fifa_ranks(wb)
    print(f"  -> {len(ranks)} selecciones cargadas.")

    ws_g  = wb["Grupos"]
    ws_ko = wb["Eliminatorias"]

    fill_group_stage(ws_g, ranks)

    print("\nCalculando clasificados...")
    standings = compute_group_standings(ws_g, ranks)
    for letter, s in standings.items():
        teams_str = ", ".join(
            f"{i+1}.{x['team']}({x['pts']}pts)" for i, x in enumerate(s)
        )
        print(f"  Grupo {letter}: {teams_str}")

    thirds = get_best_thirds(standings)
    print(f"\nMejores 8 terceros:")
    for i, t in enumerate(thirds):
        print(f"  {i+1}. {t['team']} (Gr.{t['group']}) {t['pts']}pts GD{t['gd']:+d}")

    champion, runner_up, third = fill_knockout_stage(ws_ko, standings, thirds, ranks)

    print(f"\n{'='*50}")
    print(f"  CAMPEON:       {champion}")
    print(f"  SUBCAMPEON:    {runner_up}")
    print(f"  TERCER LUGAR:  {third}")
    print(f"{'='*50}")

    wb.save(OUTPUT)
    print(f"\nArchivo guardado en:\n  {OUTPUT}")


if __name__ == "__main__":
    main()
