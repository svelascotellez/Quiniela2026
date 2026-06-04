import openpyxl

def main():
    wb_path = r"C:\Users\salva\Downloads\Quiniela_Mundial_2026_Final_vacia (1).xlsx"
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        print("Workbook loaded successfully.")
        print("Sheet names:", wb.sheetnames)
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"Sheet '{name}' dimensions: {ws.dimensions}")
    except Exception as e:
        print("Error:", e)

if __name__ == "__main__":
    main()
