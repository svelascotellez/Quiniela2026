import openpyxl

def main():
    wb_path = r"C:\Users\salva\Downloads\Quiniela_Mundial_2026_Final_vacia (1).xlsx"
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    
    # Let's inspect 'Grupos'
    ws_grupos = wb['Grupos']
    print("--- 'Grupos' sample rows (rows 5 to 15) ---")
    for r in range(5, 16):
        row_vals = [ws_grupos.cell(row=r, column=c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")

    # Let's inspect 'Ranking FIFA'
    ws_rf = wb['Ranking FIFA']
    print("\n--- 'Ranking FIFA' sample rows (rows 1 to 5) ---")
    for r in range(1, 6):
        row_vals = [ws_rf.cell(row=r, column=c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")

if __name__ == "__main__":
    main()
