import openpyxl

def main():
    wb_path = r"C:\Users\salva\Downloads\Quiniela_Mundial_2026_Final_vacia (1).xlsx"
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    ws_ko = wb['Eliminatorias']
    
    print("--- 'Eliminatorias' Dieciseisavos (Row 6-21) ---")
    for r in range(6, 22):
        t1_formula = ws_ko.cell(row=r, column=2).value
        t2_formula = ws_ko.cell(row=r, column=8).value
        # also print column 1 (Match label) and column 9 (Winner formula)
        m_label = ws_ko.cell(row=r, column=1).value
        winner_form = ws_ko.cell(row=r, column=9).value
        print(f"Row {r} ({m_label}): Col B: {t1_formula} | Col H: {t2_formula} | Col I: {winner_form}")

if __name__ == "__main__":
    main()
