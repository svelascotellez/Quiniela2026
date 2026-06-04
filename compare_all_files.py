import openpyxl
import os

def get_group_scores(path):
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Grupos']
    scores = {}
    for g in range(12):
        start_row = 6 + (8 * g)
        for m in range(6):
            r = start_row + m
            t1 = ws.cell(row=r, column=3).value
            s1 = ws.cell(row=r, column=4).value
            s2 = ws.cell(row=r, column=6).value
            t2 = ws.cell(row=r, column=7).value
            if t1 and t2:
                scores[(t1, t2)] = (s1, s2)
    return scores

def main():
    folder = r"C:\Users\salva\Downloads"
    chava = get_group_scores(os.path.join(folder, "Quiniela_Mundial_2026_Chava.xlsx"))
    salvador = get_group_scores(os.path.join(folder, "Quiniela_Mundial_2026_Salvador_Velasco.xlsx"))
    control = get_group_scores(os.path.join(folder, "Quiniela_Mundial_2026_Control.xlsx"))
    
    print("Comparing Chava and Salvador...")
    diff_chava_salvador = 0
    for k, v in chava.items():
        v2 = salvador.get(k)
        if v != v2:
            diff_chava_salvador += 1
            print(f"Diff for {k}: Chava={v}, Salvador={v2}")
    print(f"Total diffs Chava vs Salvador: {diff_chava_salvador}")
    
    print("\nComparing Chava and Control (where Control is not None)...")
    diff_control = 0
    for k, v in control.items():
        if v[0] is not None and v[1] is not None:
            v_chava = chava.get(k)
            if v != v_chava:
                diff_control += 1
                print(f"Diff for {k}: Control={v}, Chava={v_chava}")
    print(f"Total diffs Control vs Chava: {diff_control}")

if __name__ == "__main__":
    main()
