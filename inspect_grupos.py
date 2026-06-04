import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\salva\Downloads\Quiniela_Mundial_2026_Final_vacia (1).xlsx', data_only=False)
print('Hojas:', wb.sheetnames)

ws = wb['Grupos']
print('--- Filas 1-100 de Grupos (cols 1-10) ---')
for r in range(1, 110):
    row_vals = []
    has_data = False
    for c in range(1, 11):
        v = ws.cell(row=r, column=c).value
        sv = str(v) if v is not None else '_'
        row_vals.append(sv[:20])
        if v is not None:
            has_data = True
    if has_data:
        print(f'  R{r:03d}: {" | ".join(row_vals)}')
