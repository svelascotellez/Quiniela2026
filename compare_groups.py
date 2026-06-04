import openpyxl

def main():
    wb_vac = openpyxl.load_workbook(r"C:\Users\salva\Downloads\Quiniela_Mundial_2026_Final_vacia (1).xlsx", data_only=True)
    wb_ctrl = openpyxl.load_workbook(r"C:\Users\salva\Downloads\Quiniela_Mundial_2026_Control.xlsx", data_only=True)
    wb_salv = openpyxl.load_workbook(r"C:\Users\salva\Downloads\Quiniela_Mundial_2026_Salvador_Velasco.xlsx", data_only=True)
    
    ws_vac = wb_vac['Grupos']
    ws_ctrl = wb_ctrl['Grupos']
    ws_salv = wb_salv['Grupos']
    
    mismatches_ctrl = 0
    mismatches_salv = 0
    
    for g in range(12):
        start_row = 6 + (8 * g)
        for m in range(6):
            r = start_row + m
            t1_vac = ws_vac.cell(row=r, column=3).value
            t2_vac = ws_vac.cell(row=r, column=7).value
            
            t1_ctrl = ws_ctrl.cell(row=r, column=3).value
            t2_ctrl = ws_ctrl.cell(row=r, column=7).value
            
            t1_salv = ws_salv.cell(row=r, column=3).value
            t2_salv = ws_salv.cell(row=r, column=7).value
            
            if (t1_vac != t1_ctrl) or (t2_vac != t2_ctrl):
                mismatches_ctrl += 1
                print(f"Row {r} mismatch vac vs ctrl: ({t1_vac} vs {t2_vac}) vs ({t1_ctrl} vs {t2_ctrl})")
                
            if (t1_vac != t1_salv) or (t2_vac != t2_salv):
                mismatches_salv += 1
                print(f"Row {r} mismatch vac vs salv: ({t1_vac} vs {t2_vac}) vs ({t1_salv} vs {t2_salv})")
                
    print(f"Total mismatches vs Control: {mismatches_ctrl}")
    print(f"Total mismatches vs Salvador: {mismatches_salv}")

if __name__ == "__main__":
    main()
