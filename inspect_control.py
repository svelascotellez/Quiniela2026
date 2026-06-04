import openpyxl

def main():
    wb_path = r"C:\Users\salva\Downloads\Quiniela_Mundial_2026_Control.xlsx"
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb['Grupos']
    
    with open("C:/Users/salva/.antigravity/Quiniela2026/control_sheet_info.txt", "w", encoding="utf-8") as f:
        f.write("Matches in Quiniela_Mundial_2026_Control.xlsx:\n")
        for g in range(12):
            start_row = 6 + (8 * g)
            f.write(f"\nGrupo {chr(65+g)}:\n")
            for m in range(6):
                r = start_row + m
                t1 = ws.cell(row=r, column=3).value
                s1 = ws.cell(row=r, column=4).value
                s2 = ws.cell(row=r, column=6).value
                t2 = ws.cell(row=r, column=7).value
                f.write(f"  Row {r}: {t1} ({s1}) vs ({s2}) {t2}\n")

if __name__ == "__main__":
    main()
