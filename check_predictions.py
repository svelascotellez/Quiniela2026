import openpyxl
import os

def check_file(path):
    if not os.path.exists(path):
        print(f"Path does not exist: {path}")
        return
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        # Check Grupos scores
        ws_grupos = wb['Grupos']
        filled_grupos = 0
        for g in range(12):
            start_row = 6 + (8 * g)
            for m in range(6):
                r = start_row + m
                s1 = ws_grupos.cell(row=r, column=4).value
                s2 = ws_grupos.cell(row=r, column=6).value
                if s1 is not None and s2 is not None:
                    filled_grupos += 1
                    
        # Check Eliminatorias scores
        ws_ko = wb['Eliminatorias']
        filled_ko = 0
        ko_rows = list(range(6, 22)) + list(range(25, 33)) + list(range(37, 41)) + list(range(45, 47)) + [51, 56]
        for r in ko_rows:
            s1 = ws_ko.cell(row=r, column=3).value
            # Check Column G or E
            s2 = ws_ko.cell(row=r, column=7).value
            if s2 is None or str(s2).strip() == "":
                s2 = ws_ko.cell(row=r, column=5).value
            if s1 is not None and s2 is not None and s1 != "" and s2 != "":
                filled_ko += 1
                
        print(f"File: {os.path.basename(path)}")
        print(f"  Filled Grupos matches: {filled_grupos} / 72")
        print(f"  Filled Eliminatorias matches: {filled_ko}")
    except Exception as e:
        print(f"Error checking {os.path.basename(path)}: {e}")

def main():
    folder = r"C:\Users\salva\Downloads"
    files = [
        "Quiniela_Mundial_2026_Chava.xlsx",
        "Quiniela_Mundial_2026_Control.xlsx",
        "Quiniela_Mundial_2026_Salvador_Velasco.xlsx",
        "Quiniela_Mundial_2026_Totalmente_Corregida (1).xlsx",
        "Quiniela_Mundial_2026_Totalmente_Corregida.xlsx",
        "Quiniela_Mundial_2026_Final_vacia.xlsx",
        "Quiniela_Mundial_2026_Final_vacia (1).xlsx"
    ]
    for f in files:
        check_file(os.path.join(folder, f))

if __name__ == "__main__":
    main()
